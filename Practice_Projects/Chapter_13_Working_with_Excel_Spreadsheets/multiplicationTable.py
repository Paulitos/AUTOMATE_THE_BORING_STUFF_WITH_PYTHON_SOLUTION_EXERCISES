# Multiplication Table Maker Page 326 Chatper 13
import sys
import openpyxl
from openpyxl.styles import Font

def create_multiplication_table(N):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set labels in bold
    for i in range(1, N + 1):
        cell = sheet.cell(row=1, column=i + 1)
        cell.value = i
        cell.font = Font(bold=True)
        cell = sheet.cell(row=i + 1, column=1)
        cell.value = i
        cell.font = Font(bold=True)

    # Fill in multiplication results
    for i in range(1, N + 1):
        for j in range(1, N + 1):
            sheet.cell(row=i + 1, column=j + 1).value = i * j

    # Save the workbook to a file
    wb.save(f'multiplication_table_{N}x{N}.xlsx')

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python multiplicationTable.py N")
    else:
        try:
            N = int(sys.argv[1])
            create_multiplication_table(N)
            print(f"Multiplication table for {N}x{N} created successfully.")
        except ValueError:
            print("N must be an integer.")
