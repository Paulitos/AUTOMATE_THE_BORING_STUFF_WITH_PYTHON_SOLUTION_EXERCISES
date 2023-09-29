# Blank Row Inserter Chapter 13 Page 327
import openpyxl
import sys

# Check if the correct number of command line arguments is provided
if len(sys.argv) != 4:
    print("Usage: python blankRowInserter.py N M filename.xlsx")
    sys.exit(1)

# Parse command line arguments
N = int(sys.argv[1])
M = int(sys.argv[2])
filename = sys.argv[3]

# Load the Excel workbook
try:
    wb = openpyxl.load_workbook(filename)
except FileNotFoundError:
    print(f"Error: File '{filename}' not found.")
    sys.exit(1)

# Select the first sheet in the workbook
sheet = wb.active

# Create a new workbook for the modified data
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Copy the first N rows to the new workbook
for row in sheet.iter_rows(min_row=1, max_row=N):
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

# Shift the remaining rows down by M
for row in sheet.iter_rows(min_row=N + 1):
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row + M, column=cell.column, value=cell.value)

# Save the modified workbook to a new file
new_filename = f"modified_{filename}"
new_wb.save(new_filename)
print(f"Blank rows inserted successfully. Saved as {new_filename}")

