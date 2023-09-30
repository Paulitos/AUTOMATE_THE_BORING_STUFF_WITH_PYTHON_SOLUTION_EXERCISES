# Spreadsheet Cell Inverter Page 327 Chapter 13
import openpyxl

# Input the filename from the user
input_filename = input("Enter the name of the Excel file (e.g., 'your_spreadsheet.xlsx'): ")

try:
    # Load the Excel workbook
    wb = openpyxl.load_workbook(input_filename)

    # Select the active sheet (you can change this to a specific sheet if needed)
    sheet = wb.active

    # Read data from the spreadsheet into a list of lists
    sheetData = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        rowData = []
        for cell in row:
            rowData.append(cell.value)
        sheetData.append(rowData)

    # Generate the new filename
    new_filename = input_filename.replace('.xlsx', '_inverted_spreadsheet.xlsx')

    # Create a new workbook for the inverted data
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Write the inverted data back to the new workbook
    for y, row in enumerate(sheetData):
        for x, value in enumerate(row):
            new_sheet.cell(row=x+1, column=y+1, value=value)

    # Save the new workbook to the generated filename
    new_wb.save(new_filename)

    print(f"Inverted spreadsheet saved as '{new_filename}'")

except FileNotFoundError:
    print(f"File '{input_filename}' not found. Please make sure the file exists in the current directory.")
except Exception as e:
    print(f"An error occurred: {str(e)}")

# Close the workbooks
wb.close()
new_wb.close()
