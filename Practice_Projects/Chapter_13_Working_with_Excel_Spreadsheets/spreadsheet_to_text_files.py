# Spreadsheet to Text Files Page 328 Chapter 13
import openpyxl

# Open the Excel workbook
wb = openpyxl.load_workbook('output_textFiles_toSpreadsheet.xlsx')  # Replace with the path to your spreadsheet

# Select the active sheet (you can change this to a specific sheet if needed)
sheet = wb.active

# Determine the number of columns in the sheet
num_columns = sheet.max_column

# Loop through each column and write its content to a separate text file
for col in range(1, num_columns + 1):
    column_data = []
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col).value
        column_data.append(str(cell_value) if cell_value is not None else "")

    # Write the column data to a text file
    with open(f'column_{col}.txt', 'w') as text_file:
        text_file.write('\n'.join(column_data))

# Close the workbook
wb.close()
